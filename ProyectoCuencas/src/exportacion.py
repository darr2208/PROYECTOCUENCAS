import io
import geopandas as gpd
import pandas as pd
import tempfile
import zipfile
import os
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference

def exportar_shapefile_zip(gdf):
    with tempfile.TemporaryDirectory() as tmpdir:
        shp_path = os.path.join(tmpdir, "cuenca.shp")
        gdf.to_crs("EPSG:4326").to_file(shp_path, driver="ESRI Shapefile")
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w') as zipf:
            for ext in [".shp", ".shx", ".dbf", ".prj"]:
                file_path = os.path.join(tmpdir, f"cuenca{ext}")
                zipf.write(file_path, arcname=f"cuenca{ext}")
        zip_buffer.seek(0)
        return zip_buffer

def exportar_excel(resultados_dict):
    df = pd.DataFrame.from_dict(resultados_dict, orient='index', columns=['Valor'])
    df.reset_index(inplace=True)
    df.columns = ['Parámetro', 'Valor']

    wb = Workbook()
    ws = wb.active
    ws.title = "Parámetros Morfométricos"
    ws.append(['Parámetro', 'Valor'])

    for row in dataframe_to_rows(df, index=False, header=False):
        ws.append(row)

    chart = BarChart()
    chart.title = "Parámetros Morfométricos"
    chart.x_axis.title = "Parámetro"
    chart.y_axis.title = "Valor"
    chart.width = 20
    chart.height = 10

    data = Reference(ws, min_col=2, min_row=2, max_row=ws.max_row)
    cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=False)
    chart.set_categories(cats)
    ws.add_chart(chart, "E2")

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
